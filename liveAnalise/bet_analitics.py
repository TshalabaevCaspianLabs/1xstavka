from selenium import webdriver
from time import sleep
from openpyxl import load_workbook
import xlrd, xlwt

def chek_bet(name):
    analise = []
    with open('/Users/macbookpro/Documents/1xstavka/analytics_file.txt', 'r') as file:
        for gameinfo in file:
            gameinfo = gameinfo.strip()
            one, two, bet, set_ = gameinfo.split(' | ')
            analise.append([f'{one} - {two} {set_.upper()}', bet])
        file.close()

    for bet_ in analise:
        if name == bet_[0]:
            return bet_[1]



def get_analitics(driver):
    driver.get('https://1xstavka.ru/en/office/history/')
    sleep(3)
    driver.find_element_by_xpath('//*[@id="sports_page"]/div[2]/div/div/div/div[2]/div[2]/div[2]/span[2]').click()
    sleep(3)
    fulldata = []
    section_list = driver.find_elements_by_tag_name('section')

    for section in section_list:
        data = []
        p_list = section.find_elements_by_tag_name('p')

        for p in p_list:
            if p.get_attribute('class') == 'apm-panel-head__subtext':
                data.append(p.text)

            if p.get_attribute('class') == 'apm-panel-head__text':
                data.append(p.text)


        print(data[2])
        fulldata.append(data)


    with open('analise_history_bet.txt', 'w') as filename:
        for info in fulldata:
            try:
                filename.write(f'{info[0]} | {info[2]} | {info[10]} ' + '\n')
            except:
                filename.write(f'{info[0]} | {info[2]} | Loss ' + '\n')

        filename.close()


    history_bet = []
    analise = []
    game = []


    with open('/Users/macbookpro/Documents/1xstavka/analise_history_bet.txt', 'r') as filename:
        for bet in filename:
            bet = bet.strip()
            date, time, namegame, status = bet.split(' | ')
            history_bet.append([f'{date}', f'{time}', f'{namegame.upper()}', f'{status}'])
        filename.close()

    with open('/Users/macbookpro/Documents/1xstavka/analytics_file.txt', 'r') as file:
        for gameinfo in file:
            gameinfo = gameinfo.strip()
            one, two, bet, set_ = gameinfo.split(' | ')
            analise.append([f'{one} - {two} {set_.upper()}', bet])
        file.close()

    for info in analise:
        game.append(info[0])

    w = load_workbook('Analise.xlsx', use_iterators=True)
    sheet = w.worksheets[0]

    wb = xlwt.Workbook()
    ws = wb.add_sheet('Test')

    with open('analis.txt', 'w') as file_analise:
        for hbet in history_bet:
            if hbet[2] in game:
                bet_ = chek_bet(hbet[2])
                file_analise.write(f'{hbet[1]} | {hbet[2]} | {bet_} | {hbet[3]}' + '\n')

                for i in range(3):
                    row_count = sheet.max_row
                    ws.write(row_count+1, )



        file_analise.close()
